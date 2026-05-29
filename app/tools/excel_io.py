"""Read tables from an Excel file. Pure file I/O — no node logic."""
from __future__ import annotations
from io import BytesIO
from typing import Any

import openpyxl


def read_tables_from_xlsx(file_bytes: bytes) -> dict[str, list[list[Any]]]:
    """Return a dict of {sheet_name: rows} where rows is a list of lists.

    Each row is the cells in that row as Python values. Empty trailing
    rows/cols are trimmed. The first row is NOT treated as a header — the
    caller decides how to interpret the data."""
    wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    result: dict[str, list[list[Any]]] = {}
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        rows: list[list[Any]] = []
        for row in ws.iter_rows(values_only=True):
            if any(cell is not None and cell != "" for cell in row):
                rows.append(list(row))
        result[sheet] = rows
    wb.close()
    return result